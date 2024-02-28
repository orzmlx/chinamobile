import os
import pathlib
from tqdm import tqdm
import time
import pandas as pd
from param_selector import param_selector
from huaweirawdatareader import HuaweiRawDataFile
from tqdm import trange
from reporter import reporter
from zterawdatareader import ZteRawDataReader
import logging
import huaweiutils
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill
import polars as pl

def read_raw_data(path, system, date, manufacturer):
    directory = os.path.join(path, manufacturer, date, system, 'raw_data')
    if not os.path.exists(directory):
        os.makedirs(directory)
    items = pathlib.Path(directory).rglob('*')
    command_file_path = g5_command_path if system == '5G' else g4_command_path
    outputPath = os.path.join(path, manufacturer, date)
    # progress_bar = tqdm(total=len(list(items)), desc='解析【'+ manufacturer + "】原始数据...", unit='item')
    for item in items:
        if manufacturer == 'huawei':
            reader = HuaweiRawDataFile(str(item), command_file_path, outputPath, system)
            reader.read_huawei_txt()
            reader.output_format_data()
        elif manufacturer == 'zte':
            zte_config = os.path.join(path, manufacturer, '中兴网管算法_修改(2).xlsx')
            reader = ZteRawDataReader(item, outputPath, zte_config, system)
            reader.output_format_data()
        elif manufacturer == 'ericsson':
            pass
        else:
            raise Exception('未知厂商:' + manufacturer)
        del reader
        # progress_bar.update()


def combine_frq_evaluation(dir0, result_path, suffix):
    items = pathlib.Path(dir0).rglob('*')
    all_feq_path = []
    all_result = pd.DataFrame()
    for item in items:
        path = str(item)
        if path.endswith(suffix):
            all_feq_path.append(path)


def create_header(path, class_dict, base_cols):
    wb = load_workbook(path)
    sheet = wb.active
    sheet.insert_rows(1)
    base_col_len = len(base_cols)
    start_row = 1
    start_col = base_col_len + 1
    class_colors = ['48D1CC', '00FA9A', 'FFA500', '1E90FF', '800080']
    for i, clzz in enumerate(class_dict.keys()):
        clzz_cols = class_dict[clzz]
        end_column = start_col + len(clzz_cols) - 1
        end_row = start_row
        j = i
        if i >= len(class_colors):
            j = 0
        color = class_colors[j]
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row,
                          end_column=end_column)
        cell = sheet.cell(row=start_row, column=start_col, value=clzz)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(u'微软雅黑', size=10, bold=True, color='00000000')
        cell.fill = PatternFill('solid', fgColor=color)
        start_col = end_column + 1
    # 获取第二行
    header_row = sheet['2']
    for cell in header_row:
        cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='left')
        cell.fill = PatternFill('solid', fgColor='E1FFFF')
        cell.font = Font(u'微软雅黑', size=9, bold=True, color='00000000')
    wb.save(path)


def combine_evaluation(dir0, result_path, suffix, cell_class_dict):
    logging.info(">>>>>合并核查结果....<<<<<")
    items = pathlib.Path(dir0).rglob('*')
    all_result = pd.DataFrame()
    for item in items:
        path = str(item)
        if path.endswith(suffix):
            res = pd.read_csv(path)
            all_result = res if all_result.empty else pd.concat([all_result, res], axis=0)
    # all_cell_check_result_path = os.path.join(dir, check_result_name)
    all_result.to_excel(result_path, index=False, encoding='utf_8_sig',engine='openpyxl')
    huaweiutils.create_header(result_path, cell_class_dict, base_cols)


def evaluate(filepath, system, manufacturer, date, base_cols):
    raw_files = os.listdir(os.path.join(filepath, manufacturer, date, system, 'raw_data'))
    used_command = []
    cell_class_dict = {}
    freq_class_dict = {}
    for f in raw_files:
        f_name = os.path.split(f)[1]
        logging.info('>>>>>>开始处理文件:' + f_name + '<<<<<<<<')
        raw_file_dir = os.path.join(filepath, manufacturer, date, system, f_name.split('.')[0])
        if len(used_command) == 0:
            raw_fs = huaweiutils.find_file(os.path.join(raw_file_dir, 'raw_result'), '.csv')
            for f in raw_fs:
                df = pd.read_csv(f, nrows=10)
                if not df.empty:
                    command = os.path.split(f)[1].split('.')[0]
                    used_command.append(command)
        selector = param_selector(raw_file_dir, standard_path, g4_common_table, g5_common_table,
                                  g4_site_info, g5_site_info, system, used_command)
        cell_class_dict, freq_class_dict = selector.generate_report('cell', base_cols)
        del selector
    return cell_class_dict, freq_class_dict


if __name__ == '__main__':
    base_cols = ['地市', '网元', 'NRDU小区名称', 'NR小区标识', 'CGI', '频段', '工作频段',
                 '双工模式', '厂家', '共址类型', '覆盖类型', '覆盖场景', '区域类别']
    check_result_name = "all_cell_check_result.xlsx"
    cell_check_result_name = "param_check_cell.csv"
    date = '20240226'
    path = "C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\"
    g5_command_path = "C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\huawei\\20240121\\5G\\华为45G互操作固定通报参数20231225.txt"
    g4_command_path = "C:\\Users\\No.1\\Desktop\\teleccom\\华为4G异频异系统切换重选语音数据-全量.txt"
    # standard_path = "C:\\Users\\No.1\\Desktop\\teleccom\\互操作参数核查结果_test.xlsx"
    standard_path = "C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\huawei\\地市规则\\参数核查规则0223.xlsx"
    g5_common_table = "C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\huawei\\地市规则\\5G资源大表-20240131.csv"
    g4_common_table = "C:\\Users\\No.1\\Desktop\\teleccom\\LTE资源大表-0121\\LTE资源大表-0121.csv"
    g4_site_info = "C:\\Users\\No.1\\Desktop\\teleccom\\物理站CGI_4g.csv"
    g5_site_info = "C:\\Users\\No.1\\Desktop\\teleccom\\物理站CGI_5g.csv"
    raw_data_path = "C:\\Users\\No.1\\Downloads\\pytorch\\pytorch\\huawei\\result\\raw_data"
    # read_raw_data(path, '5G', '20240226', 'huawei')
    #read_raw_data(path, '5G', '20240121', 'zte')
    target_directory = os.path.join(path, 'huawei', date, '5G')
    all_cell_check_result_path = os.path.join(target_directory, check_result_name)
    report_path = os.path.join(target_directory, 'huawei', date, '互操作参数核查结果.xlsx')
    cell_class_dict, freq_class_dict = evaluate(path, '5G', 'huawei', date, base_cols)
   # combine_evaluation(target_directory, all_cell_check_result_path, cell_check_result_name, cell_class_dict)
    # freq_cell_suffix = ['LST NRCELLFREQRELATION_freq.csv', 'LST NRCELLEUTRANNFREQ_freq.csv']
    # for f in freq_cell_suffix:
    #     path = os.path.join(target_directory, f)
    #     combine_evaluation(target_directory, path, f)
    # cities = ['湖州', '杭州', '金华', '嘉兴', '丽水', '宁波', '衢州', '绍兴', '台州', '温州', '舟山', '汇总']
    # reporter = reporter(all_cell_check_result_path, path, standard_path, cities, date)
    # reporter.output_general_check_result()
